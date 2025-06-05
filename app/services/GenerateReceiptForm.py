# app/services/GenerateReceiptForm.py 수정

from sqlalchemy import text
from app.core.database import SessionLocal
from app.models.models import User
import openpyxl
import shutil
import os
from openpyxl.styles import Alignment
from datetime import datetime

def get_matched_name_and_payback(user_id):
    TEMPLATE_PATH = "/Users/gimdonghun/Downloads/수령증양식.xlsx"
    OUTPUT_DIR = "/Users/gimdonghun/Downloads/수령증_완성본"
    
    # 기존 디렉토리가 있다면 삭제
    if os.path.exists(OUTPUT_DIR):
        shutil.rmtree(OUTPUT_DIR)
    
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    with SessionLocal() as session:
        # 사용자의 면세점 타입을 동적으로 감지
        duty_free_type = detect_duty_free_type(user_id, session)
        print(f"사용자 {user_id}의 면세점 타입: {duty_free_type}")
        
        if duty_free_type == "lotte":
            # 롯데 면세점 데이터 조회 (동적 테이블)
            sql1 = """
            SELECT e.name, e."PayBack"
            FROM lotte_excel_data e
            JOIN receipt_match_log m
            ON e."receiptNumber" = m.receipt_number
            WHERE m.is_matched = TRUE AND m.user_id = :user_id;
            """
        else:
            # 신라 면세점 데이터 조회 (동적 테이블)
            sql1 = """
            SELECT e.name, e."PayBack"
            FROM shilla_excel_data e
            JOIN receipt_match_log m
            ON e."receiptNumber"::text = m.receipt_number
            WHERE m.is_matched = TRUE AND m.user_id = :user_id;
            """

        sql2 = """
        SELECT p.passport_number, p.birthday, p.name
        FROM passports p
        WHERE p.name = :name AND p.user_id = :user_id;
        """

        try:
            results = session.execute(text(sql1), {"user_id": user_id}).fetchall()
            print(f"매칭된 엑셀 데이터 조회 결과: {len(results)}건")
        except Exception as e:
            print(f"엑셀 데이터 조회 오류: {e}")
            # 테이블이 없는 경우 빈 결과 반환
            results = []
            
        printed_people = set()

        for name, payback in results:
            try:
                passport_result = session.execute(
                    text(sql2), {"name": name, "user_id": user_id}
                ).fetchone()

                if passport_result:
                    passport_number, birthday, passport_name = passport_result
                    person = (passport_name, payback, passport_number, birthday)

                    if person in printed_people:
                        continue
                    printed_people.add(person)

                    output_path = os.path.join(OUTPUT_DIR, f"{passport_name}_수령증.xlsx")
                    shutil.copy(TEMPLATE_PATH, output_path)

                    wb = openpyxl.load_workbook(output_path)
                    ws = wb.active
                    alignment = Alignment(horizontal="center", vertical="center")
                    ws["D7"] = passport_name
                    ws["D7"].alignment = alignment
                    ws["D8"] = passport_number
                    ws["D8"].alignment = alignment
                    ws["D9"] = birthday
                    ws["D9"].alignment = alignment
                    ws["D10"] = payback
                    ws["D10"].alignment = alignment
                    today = datetime.today()
                    formatted_date = f"{today.year}년    {today.month:02}월    {today.day:02}일"
                    ws["B16"] = formatted_date
                    ws["B16"].alignment = alignment
                    wb.save(output_path)
                    
                    print(f"수령증 생성 완료: {passport_name}")
                else:
                    print(f"여권 정보를 찾을 수 없음: {name}")
            except Exception as e:
                print(f"개별 수령증 생성 오류: {e}")
                continue
                
    print(f"총 {len(printed_people)}개의 수령증 생성 완료")
    return OUTPUT_DIR

def detect_duty_free_type(user_id, session):
    """사용자의 현재 데이터를 기반으로 면세점 타입을 감지"""
    try:
        # 신라 데이터 개수 확인
        shilla_count_sql = text("SELECT COUNT(*) FROM shilla_receipts WHERE user_id = :user_id")
        shilla_count = session.execute(shilla_count_sql, {"user_id": user_id}).scalar() or 0
        
        # 롯데 데이터 개수 확인
        lotte_count_sql = text("SELECT COUNT(*) FROM receipts WHERE user_id = :user_id")
        lotte_count = session.execute(lotte_count_sql, {"user_id": user_id}).scalar() or 0
        
        print(f"데이터 감지: 신라={shilla_count}, 롯데={lotte_count}")
        
        # 더 많은 데이터가 있는 쪽을 선택
        if shilla_count >= lotte_count:
            return "shilla"
        else:
            return "lotte"
            
    except Exception as e:
        print(f"면세점 타입 감지 오류: {e}")
        # 오류 시 기본값 반환
        return "lotte"