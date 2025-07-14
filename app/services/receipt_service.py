#!/usr/bin/env python3
"""
수령증 생성 및 다운로드 서비스
"""

import os
import shutil
import zipfile
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.core.database import SessionLocal
from app.models.models import ProcessingHistory
from app.core.config import settings
import tempfile

class ReceiptService:
    """수령증 생성 및 다운로드 서비스"""
    
    def __init__(self):
        # 설정에서 엑셀 템플릿 경로 가져오기
        self.template_path = os.path.join(settings.excel_template_dir, "수령증양식.xlsx")
        
    def generate_receipt_for_customer(self, upload_id: str, customer_name: str, user_id: int) -> Optional[str]:
        """특정 고객의 수령증 생성"""
        try:
            db = SessionLocal()
            
            # 고객 데이터 조회
            customer_data = self._get_customer_data(db, upload_id, customer_name, user_id)
            if not customer_data:
                print(f"고객 데이터를 찾을 수 없습니다: {customer_name}")
                return None
            
            # 임시 디렉토리에 수령증 생성
            temp_dir = tempfile.mkdtemp()
            receipt_path = os.path.join(temp_dir, f"{customer_name}의 수령증.xlsx")
            
            # 수령증 생성 (user_id 추가)
            customer_data['user_id'] = user_id
            success = self._create_receipt_file(customer_data, receipt_path)
            
            db.close()
            
            if success:
                return receipt_path
            else:
                return None
            
        except Exception as e:
            print(f"고객 수령증 생성 오류: {e}")
            return None
    
    def generate_receipts_for_session(self, upload_id: str, user_id: int) -> Optional[str]:
        """세션의 모든 고객 수령증 생성 (ZIP 파일)"""
        try:
            db = SessionLocal()
            
            # 세션의 모든 고객 데이터 조회
            customers_data = self._get_session_customers_data(db, upload_id, user_id)
            if not customers_data:
                print(f"세션 데이터를 찾을 수 없습니다: {upload_id}")
                return None
            
            # 임시 디렉토리 생성
            temp_dir = tempfile.mkdtemp()
            
            # 각 고객별 수령증 생성
            receipt_files = []
            for customer_data in customers_data:
                customer_name = customer_data['excel_name']
                if customer_name:
                    # user_id 추가
                    customer_data['user_id'] = user_id
                    receipt_file = os.path.join(temp_dir, f"{customer_name}의 수령증.xlsx")
                    if self._create_receipt_file(customer_data, receipt_file):
                        receipt_files.append(receipt_file)
            
            if not receipt_files:
                print("생성된 수령증이 없습니다.")
                return None
            
            # ZIP 파일 생성
            zip_path = os.path.join(temp_dir, "수령증.zip")
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for receipt_file in receipt_files:
                    zipf.write(receipt_file, os.path.basename(receipt_file))
            
            db.close()
            return zip_path
            
        except Exception as e:
            print(f"세션 수령증 생성 오류: {e}")
            return None
    
    def generate_receipts_for_current_session(self, user_id: int) -> Optional[str]:
        """현재 활성 세션의 모든 고객 수령증 생성 (ZIP 파일) - receipt_match_log 기반"""
        try:
            db = SessionLocal()
            
            # 현재 활성 세션의 모든 고객 데이터 조회
            customers_data = self._get_current_session_customers_data(db, user_id)
            if not customers_data:
                print(f"현재 세션에서 매칭된 데이터를 찾을 수 없습니다: user_id={user_id}")
                return None
            
            # 임시 디렉토리 생성
            temp_dir = tempfile.mkdtemp()
            
            # 각 고객별 수령증 생성
            receipt_files = []
            for customer_data in customers_data:
                customer_name = customer_data['excel_name']
                if customer_name:
                    # user_id 추가
                    customer_data['user_id'] = user_id
                    receipt_file = os.path.join(temp_dir, f"{customer_name}의 수령증.xlsx")
                    if self._create_receipt_file(customer_data, receipt_file):
                        receipt_files.append(receipt_file)
            
            if not receipt_files:
                print("현재 세션에서 생성된 수령증이 없습니다.")
                return None
            
            # ZIP 파일 생성
            zip_path = os.path.join(temp_dir, "수령증_모음.zip")
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for receipt_file in receipt_files:
                    zipf.write(receipt_file, os.path.basename(receipt_file))
            
            db.close()
            return zip_path
            
        except Exception as e:
            print(f"현재 세션 수령증 생성 오류: {e}")
            return None
    
    def _get_customer_data(self, db: Session, upload_id: str, customer_name: str, user_id: int) -> Optional[Dict]:
        """특정 고객의 데이터 조회"""
        try:
            # processing_history에서 고객 데이터 조회
            query = text("""
                SELECT 
                    excel_name,
                    passport_number,
                    birthday,
                    duty_free_type,
                    upload_id
                FROM processing_history 
                WHERE upload_id = :upload_id 
                AND excel_name = :customer_name 
                AND user_id = :user_id
                AND is_matched = true
                LIMIT 1
            """)
            
            result = db.execute(query, {
                "upload_id": upload_id,
                "customer_name": customer_name,
                "user_id": user_id
            }).fetchone()
            
            if not result:
                return None
            
            # commission_total 조회
            commission_total = self._get_commission_total(db, upload_id, customer_name, user_id, result.duty_free_type)
            
            return {
                'excel_name': result.excel_name,
                'passport_number': result.passport_number,
                'birthday': result.birthday,
                'commission_total': commission_total,
                'duty_free_type': result.duty_free_type
            }
            
        except Exception as e:
            print(f"고객 데이터 조회 오류: {e}")
            return None
    
    def _get_session_customers_data(self, db: Session, upload_id: str, user_id: int) -> List[Dict]:
        """세션의 모든 고객 데이터 조회"""
        try:
            # 고객별로 그룹화하여 조회
            query = text("""
                SELECT DISTINCT
                    excel_name,
                    passport_number,
                    birthday,
                    duty_free_type
                FROM processing_history 
                WHERE upload_id = :upload_id 
                AND user_id = :user_id
                AND is_matched = true
                AND excel_name IS NOT NULL
                ORDER BY excel_name
            """)
            
            results = db.execute(query, {
                "upload_id": upload_id,
                "user_id": user_id
            }).fetchall()
            
            customers_data = []
            for result in results:
                # 각 고객의 commission_total 조회
                commission_total = self._get_commission_total(db, upload_id, result.excel_name, user_id, result.duty_free_type)
                
                customers_data.append({
                    'excel_name': result.excel_name,
                    'passport_number': result.passport_number,
                    'birthday': result.birthday,
                    'commission_total': commission_total,
                    'duty_free_type': result.duty_free_type
                })
            
            return customers_data
            
        except Exception as e:
            print(f"세션 고객 데이터 조회 오류: {e}")
            return []
    
    def _get_commission_total(self, db: Session, upload_id: str, customer_name: str, user_id: int, duty_free_type: str) -> float:
        """고객의 총 수수료 조회 - processing_history에서 직접 조회"""
        try:
            # processing_history 테이블에서 직접 수수료 합계 조회
            query = text("""
                SELECT SUM(commission_fee) as total_commission
                FROM processing_history 
                WHERE upload_id = :upload_id 
                AND excel_name = :customer_name
                AND user_id = :user_id
                AND is_matched = true
                AND commission_fee IS NOT NULL
            """)
            
            result = db.execute(query, {
                "upload_id": upload_id,
                "customer_name": customer_name,
                "user_id": user_id
            }).fetchone()
            
            if result and result.total_commission:
                return float(result.total_commission)
            else:
                print(f"수수료 데이터 없음: upload_id={upload_id}, customer={customer_name}")
                return 0.0
            
        except Exception as e:
            print(f"수수료 조회 오류: {e}")
            import traceback
            traceback.print_exc()
            return 0.0
    
    def _create_receipt_file(self, customer_data: Dict, output_path: str) -> bool:
        """수령증 파일 생성"""
        try:
            # 템플릿 파일 복사
            shutil.copy2(self.template_path, output_path)
            
            # 엑셀 파일 열기
            workbook = load_workbook(output_path)
            worksheet = workbook['수령증']
            
            # 데이터 입력
            # D7: 성명
            worksheet['D7'] = customer_data.get('excel_name', '')
            
            # D8: 여권번호
            worksheet['D8'] = customer_data.get('passport_number', '')
            
            # D9: 생년월일
            birthday = customer_data.get('birthday')
            if birthday:
                if isinstance(birthday, str):
                    worksheet['D9'] = birthday
                else:
                    worksheet['D9'] = birthday.strftime('%Y-%m-%d')
            
            # D10: 현금 지급금액
            commission_total = customer_data.get('commission_total', 0)
            worksheet['D10'] = f"{commission_total:,.0f}원"
            
            # B16:D16: 현재 날짜
            now = datetime.now()
            date_str = f"{now.year}년           {now.month:02d}월          {now.day:02d}일"
            worksheet['B16'] = date_str
            
            # B19:D19: 수령자
            customer_name = customer_data.get('excel_name', '')
            receiver_str = f"수령자(收款人签字）：    {customer_name}    （인)"
            worksheet['B19'] = receiver_str
            
            # "매출내역" 시트에 excel_data 추가
            print(f"🔍 매출내역 시트 생성 시작 - 고객: {customer_data.get('excel_name')}")
            sales_success = self._populate_sales_details_sheet(workbook, customer_data)
            print(f"📊 매출내역 시트 생성 결과: {sales_success}")
            
            # 파일 저장
            workbook.save(output_path)
            workbook.close()
            
            print(f"수령증 생성 완료: {output_path}")
            return True
            
        except Exception as e:
            print(f"수령증 파일 생성 오류: {e}")
            return False
    
    def _populate_sales_details_sheet(self, workbook, customer_data: Dict) -> bool:
        """매출내역 시트에 excel_data 정보 추가"""
        try:
            print(f"📋 매출내역 시트 처리 시작")
            print(f"🧾 고객 데이터: {customer_data}")
            
            # "매출내역" 시트가 있는지 확인, 없으면 생성
            if "매출내역" not in workbook.sheetnames:
                print("📄 매출내역 시트 새로 생성")
                sales_sheet = workbook.create_sheet("매출내역")
            else:
                print("📄 기존 매출내역 시트 사용")
                sales_sheet = workbook["매출내역"]
            
            # 기존 데이터 모두 삭제 (헤더도 포함)
            if sales_sheet.max_row > 0:
                sales_sheet.delete_rows(1, sales_sheet.max_row)
                print(f"🗑️ 기존 데이터 삭제 완료")
            
            # 데이터베이스에서 해당 고객의 매출 데이터 조회
            print("🔍 매출 데이터 조회 시작")
            sales_data = self._get_customer_sales_data(customer_data)
            print(f"📊 조회된 매출 데이터 개수: {len(sales_data) if sales_data else 0}")
            
            if not sales_data:
                print("❌ 매출 데이터가 없습니다.")
                # 빈 시트라도 헤더는 추가
                sales_sheet.cell(row=1, column=1, value="매출 데이터가 없습니다")
                return False
            
            # 헤더와 데이터 추가
            if sales_data:
                # 첫 번째 행에 헤더 추가
                headers = list(sales_data[0].keys())
                print(f"📝 헤더: {headers}")
                for col_idx, header in enumerate(headers, 1):
                    sales_sheet.cell(row=1, column=col_idx, value=header)
                
                # 데이터 행들 추가
                for row_idx, row_data in enumerate(sales_data, 2):
                    for col_idx, value in enumerate(row_data.values(), 1):
                        sales_sheet.cell(row=row_idx, column=col_idx, value=value)
                
                print(f"✅ 매출내역 시트에 헤더 1개 + 데이터 {len(sales_data)}개 행 추가 완료")
            
            return True
            
        except Exception as e:
            print(f"❌ 매출내역 시트 생성 오류: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _get_customer_sales_data(self, customer_data: Dict) -> List[Dict]:
        """고객의 매출 데이터 조회 (processing_history + excel_data)"""
        try:
            print(f"🔍 _get_customer_sales_data 시작")
            db = SessionLocal()
            duty_free_type = customer_data.get('duty_free_type', 'lotte')
            customer_name = customer_data.get('excel_name', '')
            
            print(f"🏪 면세점 타입: {duty_free_type}")
            print(f"👤 고객명: {customer_name}")
            
            # customer_data에서 user_id 가져오기
            user_id = customer_data.get('user_id')
            if not user_id:
                print("❌ user_id가 제공되지 않았습니다.")
                db.close()
                return []
            
            print(f"🆔 사용자 ID: {user_id}")
            
            # 현재 세션 또는 processing_history에서 해당 고객의 영수증 번호들과 실제 이름 조회
            # 먼저 receipt_match_log에서 조회 시도
            receipt_query = text("""
                SELECT DISTINCT 
                    rml.receipt_number,
                    COALESCE(ph.excel_name, rml.excel_name) as actual_name
                FROM receipt_match_log rml
                LEFT JOIN processing_history ph ON rml.receipt_number = ph.receipt_number 
                    AND rml.user_id = ph.user_id
                WHERE rml.excel_name = :customer_name
                AND rml.user_id = :user_id
                AND rml.is_matched = true
                AND rml.receipt_number IS NOT NULL
            """)
            
            print(f"🔍 영수증 번호 조회 쿼리 실행 중...")
            receipt_results = db.execute(receipt_query, {
                "customer_name": customer_name,
                "user_id": user_id
            }).fetchall()
            
            print(f"📋 조회된 영수증 결과 개수: {len(receipt_results)}")
            
            if not receipt_results:
                print(f"❌ receipt_match_log에서 해당 고객을 찾을 수 없음: {customer_name}")
                print(f"🔄 processing_history에서 재시도...")
                
                # processing_history에서 직접 조회 (fallback)
                history_query = text("""
                    SELECT DISTINCT 
                        receipt_number,
                        excel_name as actual_name
                    FROM processing_history 
                    WHERE excel_name = :customer_name
                    AND user_id = :user_id
                    AND is_matched = true
                    AND receipt_number IS NOT NULL
                """)
                
                receipt_results = db.execute(history_query, {
                    "customer_name": customer_name,
                    "user_id": user_id
                }).fetchall()
                
                print(f"📋 processing_history에서 조회된 결과: {len(receipt_results)}개")
                
                if not receipt_results:
                    print(f"❌ processing_history에서도 해당 고객을 찾을 수 없음: {customer_name}")
                    
                    # 디버깅용 쿼리
                    debug_query = text("""
                        SELECT excel_name, is_matched, receipt_number, user_id
                        FROM processing_history 
                        WHERE user_id = :user_id 
                        LIMIT 10
                    """)
                    debug_results = db.execute(debug_query, {"user_id": user_id}).fetchall()
                    print(f"🔍 processing_history 샘플 데이터: {[dict(row._mapping) for row in debug_results]}")
                    
                    db.close()
                    return []
            
            # 영수증 번호들 수집
            receipt_numbers = [row.receipt_number for row in receipt_results]
            actual_name = receipt_results[0].actual_name if receipt_results else customer_name
            
            print(f"🧾 고객 '{customer_name}' (실제명: '{actual_name}')의 영수증 번호들: {receipt_numbers}")
            print(f"📝 영수증 개수: {len(receipt_numbers)}")
            
            # duty_free_type이 None인 경우 자동 감지
            if not duty_free_type:
                print("🔍 duty_free_type이 None이므로 자동 감지 시도")
                duty_free_type = self._detect_duty_free_type_by_user(user_id)
                print(f"🏪 자동 감지된 면세점 타입: {duty_free_type}")
            
            # 면세점별로 다른 테이블에서 데이터 조회
            if duty_free_type and duty_free_type.lower() == 'shilla':
                print("🏪 신라 면세점 데이터 조회 시작")
                excel_table = "shilla_excel_data"
                receipt_number_column = '"receiptNumber"'
                
                # 신라는 이름이 잘리므로 영수증 번호로만 조회하고 나중에 이름을 보정
                if receipt_numbers:
                    # PostgreSQL의 ANY 연산자 대신 IN 절 사용
                    placeholders = ','.join([f"'{num}'" for num in receipt_numbers])
                    query_str = f"""
                        SELECT *
                        FROM {excel_table}
                        WHERE {receipt_number_column}::text IN ({placeholders})
                        ORDER BY {receipt_number_column}, "상품코드"
                    """
                    print(f"🔍 신라 쿼리: {query_str}")
                    query = text(query_str)
                    
                    results = db.execute(query).fetchall()
                    print(f"📊 신라 조회 결과: {len(results)}개")
                else:
                    print("❌ 영수증 번호가 없음")
                    results = []
                
            elif duty_free_type and duty_free_type.lower() == 'lotte':
                print("🏪 롯데 면세점 데이터 조회 시작")
                excel_table = "lotte_excel_data"
                receipt_number_column = '"receiptNumber"'
                name_column = 'name'
                
                if receipt_numbers:
                    # PostgreSQL의 ANY 연산자 대신 IN 절 사용
                    placeholders = ','.join([f"'{num}'" for num in receipt_numbers])
                    query_str = f"""
                        SELECT *
                        FROM {excel_table}
                        WHERE {receipt_number_column} IN ({placeholders})
                        AND {name_column} = :customer_name
                        ORDER BY {receipt_number_column}, "상품코드"
                    """
                    print(f"🔍 롯데 쿼리: {query_str}")
                    print(f"👤 검색할 고객명: {actual_name}")
                    query = text(query_str)
                    
                    results = db.execute(query, {
                        "customer_name": actual_name
                    }).fetchall()
                    print(f"📊 롯데 조회 결과: {len(results)}개")
                else:
                    print("❌ 영수증 번호가 없음")
                    results = []
            
            else:
                print(f"❌ 알 수 없는 면세점 타입: {duty_free_type}")
                results = []
            
            if not results:
                print(f"❌ excel_data에서 해당 고객의 매출 데이터를 찾을 수 없습니다: {actual_name}")
                
                # 디버깅용: excel_data 테이블에 실제로 데이터가 있는지 확인
                if duty_free_type.lower() == 'shilla':
                    debug_query = text("SELECT COUNT(*) FROM shilla_excel_data LIMIT 1")
                else:
                    debug_query = text("SELECT COUNT(*) FROM lotte_excel_data LIMIT 1")
                
                try:
                    count = db.execute(debug_query).scalar()
                    print(f"🔍 {excel_table} 테이블 총 레코드 수: {count}")
                except Exception as e:
                    print(f"❌ {excel_table} 테이블 존재하지 않음 또는 오류: {e}")
                
                db.close()
                return []
            
            # 결과를 딕셔너리 리스트로 변환 (PayBack 컬럼 제외)
            print(f"📋 결과 변환 시작, 원본 데이터 {len(results)}개")
            sales_data = []
            for i, row in enumerate(results):
                row_dict = {}
                # SQLAlchemy Row 객체를 딕셔너리로 변환
                row_mapping = dict(row._mapping)
                
                for column, value in row_mapping.items():
                    if column.lower() != 'payback':  # PayBack 컬럼 제외
                        # 신라의 경우 name 컬럼을 실제 이름으로 보정
                        if duty_free_type.lower() == 'shilla' and column.lower() == 'name':
                            row_dict[column] = actual_name
                        else:
                            row_dict[column] = value
                sales_data.append(row_dict)
                
                # 첫 번째 행만 로그 출력
                if i == 0:
                    print(f"📝 첫 번째 행 샘플 (처음 5개 컬럼): {dict(list(row_dict.items())[:5])}")
            
            db.close()
            print(f"✅ 매출 데이터 {len(sales_data)}건 변환 완료")
            return sales_data
            
        except Exception as e:
            print(f"고객 매출 데이터 조회 오류: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def _get_current_session_customers_data(self, db: Session, user_id: int) -> List[Dict]:
        """현재 활성 세션의 모든 고객 데이터 조회 - receipt_match_log 기반"""
        try:
            # 면세점 타입 확인
            duty_free_type = self._detect_duty_free_type(db, user_id)
            
            # 현재 활성 세션에서 고객별로 그룹화하여 조회 (여권 정보와 함께)
            query = text("""
                SELECT DISTINCT
                    rml.excel_name,
                    rml.passport_number,
                    rml.birthday,
                    p.name as passport_full_name,
                    :duty_free_type as duty_free_type
                FROM receipt_match_log rml
                LEFT JOIN passports p ON rml.passport_number = p.passport_number 
                    AND p.user_id = rml.user_id
                WHERE rml.user_id = :user_id
                AND rml.is_matched = true
                AND rml.excel_name IS NOT NULL
                ORDER BY COALESCE(p.name, rml.excel_name)
            """)
            
            results = db.execute(query, {
                "user_id": user_id,
                "duty_free_type": duty_free_type
            }).fetchall()
            
            customers_data = []
            for result in results:
                # 여권 풀네임이 있으면 우선 사용, 없으면 엑셀 이름 사용
                final_name = result.passport_full_name if result.passport_full_name else result.excel_name
                
                # 각 고객의 commission_total 조회 (excel_name으로 검색)
                commission_total = self._get_current_session_commission_total(db, result.excel_name, user_id)
                
                customers_data.append({
                    'excel_name': final_name,  # 여권 풀네임 우선 사용
                    'passport_number': result.passport_number,
                    'birthday': result.birthday,
                    'commission_total': commission_total,
                    'duty_free_type': duty_free_type  # 감지된 duty_free_type 사용
                })
                
                if result.passport_full_name:
                    print(f"🏷️ 수령증 이름: 엑셀 '{result.excel_name}' → 여권 풀네임 '{result.passport_full_name}'")
            
            return customers_data
            
        except Exception as e:
            print(f"현재 세션 고객 데이터 조회 오류: {e}")
            return []
    
    def _get_current_session_commission_total(self, db: Session, customer_name: str, user_id: int) -> float:
        """현재 활성 세션에서 고객의 총 수수료 조회 - receipt_match_log 기반"""
        try:
            query = text("""
                SELECT SUM(commission_fee) as total_commission
                FROM receipt_match_log 
                WHERE excel_name = :customer_name
                AND user_id = :user_id
                AND is_matched = true
                AND commission_fee IS NOT NULL
            """)
            
            result = db.execute(query, {
                "customer_name": customer_name,
                "user_id": user_id
            }).fetchone()
            
            if result and result.total_commission:
                return float(result.total_commission)
            else:
                print(f"현재 세션 수수료 데이터 없음: customer={customer_name}")
                return 0.0
            
        except Exception as e:
            print(f"현재 세션 수수료 조회 오류: {e}")
            return 0.0
    
    def _detect_duty_free_type(self, db: Session, user_id: int) -> str:
        """사용자의 면세점 타입 감지"""
        try:
            # 신라 데이터가 있는지 확인
            shilla_count = db.execute(text("SELECT COUNT(*) FROM shilla_receipts WHERE user_id = :user_id"), 
                                     {"user_id": user_id}).scalar()
            if shilla_count > 0:
                return "shilla"
            else:
                return "lotte"
        except:
            return "lotte"  # 기본값
    
    def _detect_duty_free_type_by_user(self, user_id: int) -> str:
        """사용자의 면세점 타입 감지 (DB 연결 없이)"""
        try:
            db = SessionLocal()
            # 신라 데이터가 있는지 확인
            shilla_count = db.execute(text("SELECT COUNT(*) FROM shilla_receipts WHERE user_id = :user_id"), 
                                     {"user_id": user_id}).scalar()
            db.close()
            if shilla_count > 0:
                return "shilla"
            else:
                return "lotte"
        except:
            return "lotte"  # 기본값 

    def update_excel_names_to_passport_names(self, user_id: int) -> Dict:
        """기존 receipt_match_log 데이터의 excel_name을 여권 풀네임으로 업데이트"""
        try:
            db = SessionLocal()
            
            # 여권번호가 있고 여권 정보가 존재하는 매칭 로그 조회
            update_query = text("""
                UPDATE receipt_match_log rml
                SET excel_name = p.name
                FROM passports p
                WHERE rml.passport_number = p.passport_number
                AND rml.user_id = p.user_id
                AND rml.user_id = :user_id
                AND rml.passport_number IS NOT NULL
                AND rml.passport_number != ''
                AND p.name IS NOT NULL
                AND p.name != ''
                AND rml.excel_name != p.name
            """)
            
            updated_count = db.execute(update_query, {"user_id": user_id}).rowcount
            db.commit()
            
            print(f"✅ receipt_match_log에서 {updated_count}개 레코드의 excel_name을 여권 풀네임으로 업데이트")
            
            # processing_history도 업데이트
            history_update_query = text("""
                UPDATE processing_history ph
                SET excel_name = p.name
                FROM passports p
                WHERE ph.passport_number = p.passport_number
                AND ph.user_id = p.user_id
                AND ph.user_id = :user_id
                AND ph.passport_number IS NOT NULL
                AND ph.passport_number != ''
                AND p.name IS NOT NULL
                AND p.name != ''
                AND ph.excel_name != p.name
            """)
            
            history_updated_count = db.execute(history_update_query, {"user_id": user_id}).rowcount
            db.commit()
            
            print(f"✅ processing_history에서 {history_updated_count}개 레코드의 excel_name을 여권 풀네임으로 업데이트")
            
            db.close()
            
            return {
                "success": True,
                "updated_receipt_logs": updated_count,
                "updated_history": history_updated_count,
                "message": f"총 {updated_count + history_updated_count}개 레코드 업데이트 완료"
            }
            
        except Exception as e:
            print(f"기존 데이터 업데이트 오류: {e}")
            return {
                "success": False,
                "error": str(e)
            } 