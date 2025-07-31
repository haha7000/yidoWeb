import pandas as pd
import asyncio
import aiohttp
import os
import warnings
import zipfile
import re
import subprocess
from datetime import datetime
warnings.filterwarnings('ignore')

class ExcelProcessor:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path
        self.processed_file_path = None
        
    def fix_excel_date_format(self):
        """Excel 파일의 사용자지정 날짜 형식 문제를 해결하는 함수"""
        try:
            print("Excel 날짜 형식 문제 해결 시도 중...")
            
            # 임시 파일 생성
            temp_file = self.excel_file_path.replace('.xlsx', '_fixed_dates.xlsx')
            
            # 원본 파일을 zip으로 열어서 내용 확인
            with zipfile.ZipFile(self.excel_file_path, 'r') as zip_ref:
                # 파일 목록 확인
                file_list = zip_ref.namelist()
                print(f"Excel 파일 내부 구조: {file_list}")
                
                # 새로운 zip 파일 생성
                with zipfile.ZipFile(temp_file, 'w') as new_zip:
                    for file_name in file_list:
                        try:
                            content = zip_ref.read(file_name)
                            
                            # XML 파일인 경우 날짜 형식 수정
                            if file_name.endswith('.xml'):
                                content_str = content.decode('utf-8')
                                
                                # 사용자지정 날짜 형식 문제 해결
                                # 1. 잘못된 ISO8601 형식을 올바른 형식으로 변환
                                def fix_iso8601_date(match):
                                    date_str = match.group(1)
                                    if len(date_str) == 15 and 'T' in date_str:  # 20250722T000000 형식
                                        year = date_str[:4]
                                        month = date_str[4:6]
                                        day = date_str[6:8]
                                        time = date_str[9:15]
                                        return f"{year}-{month}-{day}T{time[:2]}:{time[2:4]}:{time[4:6]}"
                                    return date_str
                                
                                # 2. 날짜 셀 타입을 문자열로 변경
                                content_str = re.sub(r'<c r="[^"]*" t="d"><v>(\d{8}T\d{6})</v></c>', 
                                                   r'<c r="\g<0>" t="s"><v>\1</v></c>', content_str)
                                
                                # 3. 일반적인 잘못된 날짜 패턴 수정
                                content_str = re.sub(r'(\d{8}T\d{6})', fix_iso8601_date, content_str)
                                
                                content = content_str.encode('utf-8')
                            
                            new_zip.writestr(file_name, content)
                        except Exception as e:
                            print(f"파일 {file_name} 처리 중 오류: {e}")
                            continue
            
            return temp_file
        except Exception as e:
            print(f"날짜 형식 수정 실패: {e}")
            return None
        
    def read_excel_with_date_fix(self):
        """날짜 형식 문제를 해결하여 Excel 파일을 읽는 함수"""
        try:
            # 1. 날짜 형식 수정
            fixed_file = self.fix_excel_date_format()
            
            if fixed_file and os.path.exists(fixed_file):
                print(f"수정된 파일 사용: {fixed_file}")
                file_to_read = fixed_file
            else:
                print("원본 파일 사용")
                file_to_read = self.excel_file_path
            
            # 2. 여러 방법으로 Excel 파일 읽기 시도
            methods = [
                lambda: pd.read_excel(file_to_read, engine='openpyxl', parse_dates=False),
                lambda: pd.read_excel(file_to_read, engine='openpyxl', parse_dates=False, na_filter=False),
                lambda: pd.read_excel(file_to_read, engine='openpyxl', parse_dates=False, keep_default_na=False),
                lambda: pd.read_excel(file_to_read, engine='openpyxl', parse_dates=False, date_parser=lambda x: pd.to_datetime(x, errors='coerce'))
            ]
            
            for i, method in enumerate(methods):
                try:
                    print(f"Excel 읽기 방법 {i+1} 시도 중...")
                    df = method()
                    print("✅ Excel 파일 읽기 성공!")
                    
                    # 수정된 파일이 있다면 삭제
                    if fixed_file and os.path.exists(fixed_file):
                        os.remove(fixed_file)
                        print(f"임시 파일 삭제: {fixed_file}")
                    
                    return df
                except Exception as e:
                    print(f"방법 {i+1} 실패: {e}")
                    continue
            
            # 3. 최종 방법: 바이너리로 직접 읽기
            print("바이너리 방식으로 시도 중...")
            try:
                with open(file_to_read, 'rb') as f:
                    from openpyxl import load_workbook
                    wb = load_workbook(f, data_only=True, read_only=True)
                    ws = wb.active
                    
                    # 데이터 추출
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            data.append(row)
                    
                    if data:
                        df = pd.DataFrame(data[1:], columns=data[0])
                        print("✅ 바이너리 방식으로 성공!")
                        
                        # 수정된 파일이 있다면 삭제
                        if fixed_file and os.path.exists(fixed_file):
                            os.remove(fixed_file)
                            print(f"임시 파일 삭제: {fixed_file}")
                        
                        return df
                    else:
                        print("❌ 데이터를 추출할 수 없습니다.")
                        
            except Exception as e:
                print(f"바이너리 방식도 실패: {e}")
            
            raise Exception("모든 Excel 읽기 방법이 실패했습니다.")
            
        except Exception as e:
            print(f"Excel 파일 읽기 실패: {e}")
            return None
    
    def rename_columns(self, df):
        """컬럼명 변경"""
        column_mapping = {
            'BILL 번호': 'receiptNumber',
            '고객명': 'name'
        }
        
        # 기존 컬럼명 확인
        print("현재 컬럼명:")
        print(df.columns.tolist())
        
        # 컬럼명 변경
        df = df.rename(columns=column_mapping)
        
        print("\n변경된 컬럼명:")
        print(df.columns.tolist())
        
        return df
    
    def convert_data_types(self, df):
        """데이터 타입 변환"""
        print("\n데이터 타입 변환 중...")
        
        # 텍스트로 변환할 컬럼들
        text_columns = [
            'No', '점', '원매출일자', '매출일자', '여행사명', '여행사코드', 
            '그룹번호', '대표가이드', '출생연도', 'name', 'BILL 상태', 
            '상품위치', '카테고리', '브랜드명', '상품명', '상품코드', 
            'REF NO', 'Aging', '판매형태', '판매수량', '총매출액($)', 
            '총매출액(￦)', '순매출액($)', '할인액($)', '', 'passport_number'
        ]
        
        # 숫자로 변환할 컬럼들
        numeric_columns = {
            '판매가($)': 'float64',
            '순매출액(￦)': 'float64', 
            '할인액(￦)': 'float64'
        }
        
        # receiptNumber는 varchar(50)로 처리
        if 'receiptNumber' in df.columns:
            df['receiptNumber'] = df['receiptNumber'].astype(str).str[:50]
        
        # 텍스트 컬럼들 변환
        for col in text_columns:
            if col in df.columns:
                df[col] = df[col].astype(str)
                print(f"컬럼 '{col}' -> text 변환 완료")
        
        # 숫자 컬럼들 변환
        for col, dtype in numeric_columns.items():
            if col in df.columns:
                try:
                    # 숫자가 아닌 값들을 NaN으로 처리
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                    df[col] = df[col].astype(dtype)
                    print(f"컬럼 '{col}' -> {dtype} 변환 완료")
                except Exception as e:
                    print(f"컬럼 '{col}' 변환 실패: {e}")
        
        return df
    
    def save_processed_excel(self, df):
        """처리된 Excel 파일 저장"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.processed_file_path = f"/Users/ec2-user/Downloads/shilla_report_processed_{timestamp}.xlsx"
        
        try:
            df.to_excel(self.processed_file_path, index=False, engine='openpyxl')
            print(f"✅ 처리된 Excel 파일 저장 완료: {self.processed_file_path}")
            return True
        except Exception as e:
            print(f"❌ Excel 파일 저장 실패: {e}")
            return False
    
    def process_excel(self):
        """Excel 파일 전체 처리 과정"""
        print("=== Excel 파일 처리 시작 ===")
        
        # 1. Excel 파일 읽기
        df = self.read_excel_with_date_fix()
        if df is None:
            return None
        
        # 2. 컬럼명 변경
        df = self.rename_columns(df)
        
        # 3. 데이터 타입 변환
        df = self.convert_data_types(df)
        
        # 4. 처리된 파일 저장
        if not self.save_processed_excel(df):
            return None
        
        print("=== Excel 파일 처리 완료 ===")
        return df

async def upload_processed_excel(file_path):
    """처리된 Excel 파일을 업로드 API로 전송"""
    try:
        print(f"📤 처리된 Excel 업로드 시작: {file_path}")
        
        # 파일 존재 확인
        if not os.path.exists(file_path):
            print(f"❌ 파일이 존재하지 않습니다: {file_path}")
            return False
        
        # 업로드 API 설정 (EC2 공인 IP 사용)
        # EC2의 공인 IP를 여기에 입력하세요
        ec2_ip = "18.142.243.40"  # EC2 공인 IP
        login_url = f"http://{ec2_ip}:8001/login/"
        api_url = f"http://{ec2_ip}:8001/upload-excel/"
        
        headers = {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
        
        # 더 강력한 세션 관리를 위한 설정
        timeout = aiohttp.ClientTimeout(total=60)
        connector = aiohttp.TCPConnector(ssl=False)
        
        async with aiohttp.ClientSession(
            cookie_jar=aiohttp.CookieJar(),
            timeout=timeout,
            connector=connector
        ) as session:
            # 로그인 페이지에서 먼저 세션 시작
            print("🔐 로그인 페이지 접근 중...")
            async with session.get(f"http://{ec2_ip}:8001/", headers=headers) as initial_resp:
                print(f"초기 페이지 상태: {initial_resp.status}")
            
            # 로그인 시도
            login_data = {
                "username": "haha",
                "password": "haha"
            }
            
            print("🔐 웹사이트 로그인 시도...")
            async with session.post(login_url, data=login_data, headers=headers, allow_redirects=False) as login_resp:
                print(f"로그인 응답 상태: {login_resp.status}")
                print(f"응답 헤더: {dict(login_resp.headers)}")
                
                if login_resp.status == 302:  # 리다이렉트 성공
                    print("✅ 로그인 성공! (리다이렉트 감지)")
                    
                    # Set-Cookie 헤더에서 직접 쿠키 추출
                    set_cookie_header = login_resp.headers.get('Set-Cookie')
                    if set_cookie_header:
                        print(f"🍪 Set-Cookie 헤더: {set_cookie_header}")
                        
                        # access_token 쿠키 추출
                        import re
                        token_match = re.search(r'access_token=([^;]+)', set_cookie_header)
                        if token_match:
                            access_token = token_match.group(1)
                            print(f"🔑 JWT 토큰 추출 성공: {access_token[:20]}...")
                            
                            # HttpOnly 쿠키 문제 해결을 위해 다른 방법 시도
                            try:
                                # 방법 1: SimpleCookie를 사용한 수동 설정
                                from http.cookies import SimpleCookie
                                from yarl import URL
                                
                                cookie = SimpleCookie()
                                cookie['access_token'] = access_token
                                cookie['access_token']['path'] = '/'
                                cookie['access_token']['domain'] = ec2_ip
                                
                                # 쿠키 jar에 직접 추가
                                import datetime
                                expires = datetime.datetime.now() + datetime.timedelta(days=1)
                                
                                session.cookie_jar._cookies.clear()  # 기존 쿠키 클리어
                                session.cookie_jar._cookies[ec2_ip] = {
                                    '/': {
                                        'access_token': aiohttp.cookiejar.Morsel(
                                            key='access_token',
                                            value=access_token,
                                            coded_value=access_token,
                                            rest={'path': '/', 'domain': ec2_ip}
                                        )
                                    }
                                }
                                print("🍪 쿠키 수동 설정 완료 (방법 1)")
                            except Exception as cookie_error:
                                print(f"⚠️ 쿠키 설정 방법 1 실패: {cookie_error}")
                                
                                # 방법 2: 헤더에 직접 쿠키 추가
                                print("🔄 쿠키를 헤더로 직접 전송하는 방식으로 변경")
                                global manual_cookie_header
                                manual_cookie_header = f"access_token={access_token}"
                                print(f"🍪 쿠키 헤더 준비: {manual_cookie_header[:50]}...")
                        else:
                            print("❌ Set-Cookie 헤더에서 access_token을 찾을 수 없습니다")
                            return False
                    else:
                        print("❌ Set-Cookie 헤더가 없습니다")
                        return False
                    
                    # 리다이렉트 위치 확인
                    redirect_location = login_resp.headers.get('Location', '/upload/')
                    print(f"리다이렉트 위치: {redirect_location}")
                    
                    # 쿠키 설정 후 최종 확인
                    cookies_after_manual = session.cookie_jar.filter_cookies(f"http://{ec2_ip}:8001")
                    print("🍪 수동 설정 후 쿠키들:")
                    for cookie in cookies_after_manual.values():
                        print(f"  - {cookie.key}: {cookie.value[:20] if len(cookie.value) > 20 else cookie.value}...")
                    
                    # 리다이렉트 페이지 테스트 (선택사항)
                    redirect_url = f"http://{ec2_ip}:8001{redirect_location}" if redirect_location.startswith('/') else redirect_location
                    async with session.get(redirect_url, headers=headers) as redirect_resp:
                        print(f"리다이렉트 페이지 상태: {redirect_resp.status}")
                        if redirect_resp.status == 401:
                            print("⚠️ 리다이렉트 페이지에서 401 오류, 하지만 업로드는 시도해봅니다")
                        elif redirect_resp.status == 200:
                            print("✅ 리다이렉트 페이지 접근 성공!")
                    
                    # 토큰이 있으면 계속 진행
                    if 'access_token' in locals():
                        print("🔑 JWT 토큰 확인됨, 업로드 진행합니다")
                        # 업로드용 헤더에 쿠키 추가
                        upload_headers = headers.copy()
                        if 'manual_cookie_header' in globals():
                            upload_headers['Cookie'] = manual_cookie_header
                            print(f"🍪 Cookie 헤더 설정: {manual_cookie_header[:50]}...")
                        else:
                            upload_headers['Cookie'] = f"access_token={access_token}"
                            print(f"🍪 Cookie 헤더 설정: access_token={access_token[:20]}...")
                    else:
                        print("❌ JWT 토큰을 찾을 수 없습니다")
                        return False
                    
                elif login_resp.status == 200:
                    # 로그인 페이지가 다시 나타남 (로그인 실패)
                    error_text = await login_resp.text()
                    print(f"❌ 로그인 실패: 인증 정보가 잘못되었습니다")
                    if "error" in error_text:
                        print("로그인 에러 메시지가 페이지에 포함되어 있습니다")
                    return False
                else:
                    print(f"❌ 로그인 실패: {login_resp.status}")
                    error_text = await login_resp.text()
                    print(f"로그인 오류: {error_text[:200]}...")
                    return False
            
            # 엑셀 파일 업로드
            print("📤 처리된 엑셀 파일 업로드 중...")
            with open(file_path, 'rb') as f:
                form_data = aiohttp.FormData()
                form_data.add_field('excel_file', f, filename=os.path.basename(file_path))
                form_data.add_field('duty_free_type', 'shilla')
                
                # 쿠키 기반 인증 사용 (JWT 토큰을 헤더로 직접 전송)
                async with session.post(api_url, data=form_data, headers=upload_headers) as resp:
                    print(f"업로드 응답 상태: {resp.status}")
                    
                    if resp.status == 200:
                        try:
                            result = await resp.json()
                            print("✅ 처리된 Excel 업로드 성공!")
                            print(f"📊 업로드 결과: {result}")
                            return True
                        except Exception:
                            # JSON이 아닐 수 있으므로 텍스트로 시도
                            result_text = await resp.text()
                            print("✅ 처리된 Excel 업로드 성공!")
                            print(f"📊 업로드 결과 (텍스트): {result_text[:200]}...")
                            return True
                    else:
                        error_text = await resp.text()
                        print(f"❌ Excel 업로드 실패: {resp.status}")
                        print(f"오류 내용: {error_text[:500]}...")
                        
                        # 401 오류인 경우 인증 문제 추가 진단
                        if resp.status == 401:
                            print("🔍 인증 문제 진단:")
                            cookies_at_upload = session.cookie_jar.filter_cookies(f"http://{ec2_ip}:8001")
                            print(f"업로드 시점 쿠키 수: {len(list(cookies_at_upload.values()))}")
                            for cookie in cookies_at_upload.values():
                                print(f"  - {cookie.key}: {'존재함' if cookie.value else '비어있음'}")
                        
                        return False
                        
    except Exception as e:
        print(f"❌ Excel 업로드 중 오류 발생: {e}")
        import traceback
        print(f"상세 오류: {traceback.format_exc()}")
        return False

async def process_and_upload_excel(excel_file_path):
    """Excel 파일을 처리하고 업로드하는 통합 함수"""
    try:
        print(f"🚀 Excel 처리 및 업로드 시작: {excel_file_path}")
        
        # 1. Excel 파일 처리
        processor = ExcelProcessor(excel_file_path)
        processed_df = processor.process_excel()
        
        if processed_df is None:
            print("❌ Excel 파일 처리 실패")
            return False
        
        # 2. 처리된 파일 업로드
        if processor.processed_file_path:
            success = await upload_processed_excel(processor.processed_file_path)
            
            if success:
                print("🎉 Excel 처리 및 업로드 완료!")
                return True
            else:
                print("❌ 업로드 실패")
                return False
        else:
            print("❌ 처리된 파일 경로가 없습니다")
            return False
            
    except Exception as e:
        print(f"❌ Excel 처리 및 업로드 중 오류 발생: {e}")
        return False

async def main():
    """메인 실행 함수 (기존 호환성 유지)"""
    # 원본 Excel 파일 경로
    original_file = '/Users/ec2-user/Downloads/shilla_report.xlsx'
    
    # 통합 함수 호출
    success = await process_and_upload_excel(original_file)
    
    if success:
        print("🎉 전체 프로세스 완료!")
    else:
        print("❌ 프로세스 실패")

if __name__ == "__main__":
    asyncio.run(main()) 
