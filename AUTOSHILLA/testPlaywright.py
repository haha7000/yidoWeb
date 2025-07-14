from datetime import datetime, timedelta
from playwright.sync_api import sync_playwright
import requests
import json
import re
import time
import pandas as pd
import os

def check_network_connection():
    """네트워크 연결 상태 확인"""
    import urllib.request
    try:
        print("🔄 네트워크 연결 확인 중...")
        # Google DNS로 연결 테스트
        urllib.request.urlopen('http://www.google.com', timeout=10)
        print("✔ 네트워크 연결 정상")
        return True
    except Exception as e:
        print(f"❌ 네트워크 연결 실패: {e}")
        return False

def automate_shilla_download():
    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=[
                '--no-sandbox',
                '--disable-web-security',
                '--disable-features=VizDisplayCompositor',
                '--disable-dev-shm-usage'
            ]
        )
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
            viewport={'width': 1920, 'height': 1080}
        )
        page = context.new_page()
        
        # 타임아웃 설정 증가
        page.set_default_timeout(60000)  # 60초

        try:
            # 1. 로그인 페이지 접근 (재시도 로직 추가)
            print("🔄 로그인 페이지 접근 중...")
            
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    print(f"  시도 {attempt + 1}/{max_retries}")
                    page.goto("https://www.shillasrm.com/login.do", wait_until="domcontentloaded", timeout=60000)
                    print("✔ 페이지 로드 성공")
                    break
                except Exception as e:
                    print(f"  페이지 로드 실패: {e}")
                    if attempt == max_retries - 1:
                        raise Exception(f"로그인 페이지 접근 실패 (최대 {max_retries}회 시도)")
                    print("  3초 후 재시도...")
                    time.sleep(3)
            
            # 페이지 완전 로드 대기
            try:
                page.wait_for_load_state("networkidle", timeout=30000)
            except:
                print("⚠️ 네트워크 idle 대기 타임아웃 (계속 진행)")
            
            time.sleep(2)
            
            # 로그인 폼 확인
            print("🔄 로그인 폼 확인 중...")
            try:
                # 로그인 폼이 나타날 때까지 대기
                page.wait_for_selector('#username', timeout=20000)
                page.wait_for_selector('#idbox', timeout=20000)
                print("✔ 로그인 폼 확인됨")
            except Exception as e:
                print(f"❌ 로그인 폼을 찾을 수 없습니다: {e}")
                print("현재 URL:", page.url)
                print("페이지 타이틀:", page.title())
                raise
            
            # 로그인 폼 입력
            print("🔄 로그인 정보 입력 중...")
            page.fill('#username', "G000056324")
            page.fill('#idbox', "19850327ng@!!")
            
            # 로그인 버튼 클릭
            print("🔄 로그인 버튼 클릭...")
            page.click('.box_login')
            
            # 로그인 후 페이지 대기 (여러 방법으로 시도)
            print("🔄 로그인 처리 중...")
            try:
                # 방법 1: URL 변경 대기
                page.wait_for_url("**/afterLoginOK.do", timeout=30000)
                print("✔ 로그인 성공 (URL 확인)")
            except:
                try:
                    # 방법 2: 특정 요소 대기
                    page.wait_for_selector("body", timeout=20000)
                    current_url = page.url
                    if "afterLoginOK.do" in current_url or "main" in current_url:
                        print("✔ 로그인 성공 (URL 패턴 확인)")
                    else:
                        print(f"⚠️ 예상과 다른 URL: {current_url}")
                except Exception as e:
                    print(f"❌ 로그인 실패: {e}")
                    raise
            
            # 추가 대기
            try:
                page.wait_for_load_state("networkidle", timeout=20000)
            except:
                print("⚠️ 네트워크 idle 대기 타임아웃 (계속 진행)")
            
            time.sleep(5)
            
            # 2. 토큰 수집 (다양한 방법으로 시도)
            print("🔄 토큰 수집 중...")
            
            # CSRF 토큰 수집 (여러 방법으로 시도)
            csrf_token = None
            csrf_header = None
            
            try:
                # 방법 1: meta 태그에서 수집
                csrf_token = page.locator('meta[name="_csrf"]').get_attribute('content')
                csrf_header = page.locator('meta[name="_csrf_parameter"]').get_attribute('content')
                print(f"Meta 태그에서 CSRF 토큰: {csrf_token}")
            except Exception as e:
                print(f"Meta 태그에서 CSRF 토큰 수집 실패: {e}")
            
            if not csrf_token:
                try:
                    # 방법 2: 페이지 소스에서 정규식으로 찾기
                    content = page.content()
                    csrf_match = re.search(r'_csrf["\']?\s*:\s*["\']([^"\']+)', content)
                    if csrf_match:
                        csrf_token = csrf_match.group(1)
                        print(f"정규식에서 CSRF 토큰: {csrf_token}")
                except Exception as e:
                    print(f"정규식에서 CSRF 토큰 수집 실패: {e}")
            
            # MBPF 토큰 수집 (여러 방법으로 시도)
            mbpf_token = None
            try:
                # 방법 1: localStorage에서 수집
                mbpf_token = page.evaluate("() => localStorage.getItem('x-mbpf-token')")
                print(f"LocalStorage에서 MBPF 토큰: {mbpf_token}")
            except Exception as e:
                print(f"LocalStorage에서 MBPF 토큰 수집 실패: {e}")
            
            if not mbpf_token:
                try:
                    # 방법 2: sessionStorage에서 수집
                    mbpf_token = page.evaluate("() => sessionStorage.getItem('x-mbpf-token')")
                    print(f"SessionStorage에서 MBPF 토큰: {mbpf_token}")
                except Exception as e:
                    print(f"SessionStorage에서 MBPF 토큰 수집 실패: {e}")
            
            if not mbpf_token:
                try:
                    # 방법 3: 페이지 소스에서 정규식으로 찾기
                    content = page.content()
                    mbpf_match = re.search(r'x-mbpf-token["\']?\s*:\s*["\']([^"\']+)', content)
                    if mbpf_match:
                        mbpf_token = mbpf_match.group(1)
                        print(f"정규식에서 MBPF 토큰: {mbpf_token}")
                except Exception as e:
                    print(f"정규식에서 MBPF 토큰 수집 실패: {e}")
            
            # 3. 페이지 내용 디버깅
            print("🔍 페이지 내용 디버깅...")
            
            # 현재 URL 확인
            current_url = page.url
            print(f"현재 URL: {current_url}")
            
            # 페이지 타이틀 확인
            title = page.title()
            print(f"페이지 타이틀: {title}")
            
            # meta 태그들 확인
            meta_tags = page.locator('meta').all()
            print(f"Meta 태그 수: {len(meta_tags)}")
            for i, meta in enumerate(meta_tags[:10]):  # 처음 10개만 확인
                try:
                    name = meta.get_attribute('name')
                    content = meta.get_attribute('content')
                    if name:
                        print(f"  Meta[{i}]: {name} = {content}")
                except:
                    pass
            
            # localStorage 내용 확인
            try:
                local_storage = page.evaluate("() => Object.keys(localStorage)")
                print(f"LocalStorage 키들: {local_storage}")
                for key in local_storage:
                    value = page.evaluate(f"() => localStorage.getItem('{key}')")
                    print(f"  {key}: {value}")
            except Exception as e:
                print(f"LocalStorage 확인 실패: {e}")
            
            # 4. 실제 데이터 요청 페이지로 이동해서 토큰 확인
            print("🔄 데이터 요청 페이지로 이동...")
            try:
                # 실제 데이터 페이지로 이동 (예: 매출 조회 페이지)
                for attempt in range(3):
                    try:
                        print(f"  페이지 이동 시도 {attempt + 1}/3")
                        page.goto("https://www.shillasrm.com/ui/sp/external/guide/guideSale.do", timeout=45000)
                        page.wait_for_load_state("domcontentloaded", timeout=30000)
                        print("✔ 데이터 페이지 이동 성공")
                        break
                    except Exception as e:
                        print(f"  데이터 페이지 이동 실패: {e}")
                        if attempt == 2:
                            print("⚠️ 데이터 페이지 이동을 포기하고 현재 페이지에서 진행")
                        else:
                            time.sleep(3)
                
                time.sleep(3)
                
                # 다시 토큰 수집 시도
                if not csrf_token:
                    csrf_token = page.locator('meta[name="_csrf"]').get_attribute('content')
                    csrf_header = page.locator('meta[name="_csrf_parameter"]').get_attribute('content')
                    print(f"데이터 페이지에서 CSRF 토큰: {csrf_token}")
                
                if not mbpf_token:
                    mbpf_token = page.evaluate("() => localStorage.getItem('x-mbpf-token')")
                    print(f"데이터 페이지에서 MBPF 토큰: {mbpf_token}")
                
            except Exception as e:
                print(f"데이터 페이지 이동 실패: {e}")
            
            # 5. 토큰 확인
            print(f"최종 CSRF 토큰: {csrf_token}")
            print(f"최종 CSRF 헤더: {csrf_header}")
            print(f"최종 MBPF 토큰: {mbpf_token}")
            
            if not csrf_token:
                print("❌ CSRF 토큰을 찾지 못했습니다.")
                
                # 브라우저를 열어둔 상태에서 수동으로 확인할 수 있도록 대기
                print("🔍 브라우저를 열어둔 상태입니다. 개발자 도구로 토큰을 확인해보세요.")
                print("  - F12 -> Console -> localStorage.getItem('x-mbpf-token')")
                print("  - F12 -> Elements -> <meta name='_csrf'> 태그 확인")
                input("계속하려면 Enter를 눌러주세요...")
                return
            
            if not mbpf_token:
                print("⚠️ MBPF 토큰을 찾지 못했습니다. CSRF 토큰만으로 시도해보겠습니다.")
                # MBPF 토큰 없이 진행해보기
            
            # 6. 쿠키 수집
            cookies = context.cookies()
            cookie_dict = {cookie['name']: cookie['value'] for cookie in cookies}
            cookie_header = "; ".join([f"{c['name']}={c['value']}" for c in cookies])
            
            print(f"🍪 쿠키 수: {len(cookies)}")
            
            # 7. 헤더 구성
            headers = {
                "accept": "application/json",
                "content-type": "application/json",
                "origin": "https://www.shillasrm.com",
                "referer": "https://www.shillasrm.com/ui/sp/external/guide/guideSale.do",
                "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
                "x-csrf-token": csrf_token,
                "x-requested-with": "XMLHttpRequest",
                "cookie": cookie_header,
                "menucode": "DEG00020"
            }
            
            # MBPF 토큰이 있는 경우에만 추가
            if mbpf_token:
                headers["x-mbpf-token"] = mbpf_token
                print(f"✔ MBPF 토큰 추가됨: {mbpf_token}")
            else:
                print("⚠️ MBPF 토큰 없이 진행")
            
            # 8. 날짜 구성
            yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y%m%d")
            payload = {
                "sale_dt_from": yesterday,
                "sale_dt_to": yesterday,
                "store": "01",
                "grp_no_from": "",
                "grp_no_to": ""
            }
            
            print(f"📅 조회 날짜: {yesterday}")
            print(f"📊 요청 데이터: {payload}")
            
            # 9. API 호출
            print("🔄 API 호출 중...")
            session = requests.Session()
            response = session.post(
                "https://www.shillasrm.com/ui/sp/external/guide/findListGuideSale.do",
                headers=headers,
                data=json.dumps(payload)
            )
            
            print(f"✔ API 응답 상태: {response.status_code}")
            print(f"✔ API 응답 헤더: {dict(response.headers)}")
            
            try:
                json_response = response.json()
                print(f"✔ API 응답 성공! 데이터 수: {len(json_response.get('BODY', {}).get('t_result', []))}")
                
                # 데이터 저장
                save_data_to_file(json_response, yesterday)
                save_data_to_excel(json_response, yesterday)
                
                # 데이터 요약 출력
                print_data_summary(json_response)
                
                return json_response
                
            except Exception as e:
                print(f"❌ 응답을 JSON으로 파싱하지 못했습니다: {e}")
                print(f"응답 텍스트: {response.text}")
                return None
            
        except Exception as e:
            print(f"❌ 전체 프로세스 실패: {e}")
            import traceback
            traceback.print_exc()
        
        finally:
            # 브라우저 닫기 (디버깅 시에는 주석 처리)
            try:
                browser.close()
            except:
                pass

def save_data_to_file(data, date_str):
    """데이터를 Excel 파일로 저장 (한국어 컬럼명)"""
    try:
        results = data.get('BODY', {}).get('t_result', [])
        if not results:
            print("❌ 저장할 데이터가 없습니다.")
            return
        
        # 컬럼명 매핑 (영어 -> 한국어)
        column_mapping = {
            "travel_cd": "여행사코드",
            "bill_no": "BILL 번호",
            "prdt_loc_fg_nm": "인도장",
            "brand_nm": "브랜드명",
            "dc_amt": "할인액($)",
            "guide_nm": "대표가이드",
            "cust_nm_eng": "고객명",
            "price": "판매가($)",
            "travel_nm_lcl": "여행사명",
            "prdt_cd": "상품코드",
            "tot_wamt": "총매출액(￦)",
            "sale_dt": "매출일자",
            "dc_wamt": "할인액(￦)",
            "grp_no": "그룹번호",
            "net_wamt": "순매출액(￦)",
            "ref_no": "REF NO",
            "store": "점",
            "guide_resi_no": "출생연도",
            "tot_amt": "총매출액($)",
            "sale_fg_nm": "판매형태",
            "net_amt": "순매출액($)",
            "bill_proc_fg_nm": "BILL 상태",
            "cat_nm": "카테고리",
            "prdt_nm": "상품명",
            "qty": "판매수량",
            "org_sale_tr_dt": "원매출일자",
            "aging": "Aging"
        }
        
        # DataFrame 생성
        df = pd.DataFrame(results)
        
        # 컬럼명 변경
        df = df.rename(columns=column_mapping)
        
        # 컬럼 순서 정리 (중요한 컬럼들을 앞쪽으로)
        priority_columns = [
            "매출일자", "대표가이드", "여행사명", "고객명", 
            "브랜드명", "상품명", "카테고리", "판매수량",
            "판매가($)", "순매출액($)", "할인액($)",
            "총매출액(￦)", "순매출액(￦)", "할인액(￦)",
            "판매형태", "BILL 상태", "인도장"
        ]
        
        # 우선순위 컬럼 + 나머지 컬럼 순서로 정렬
        remaining_columns = [col for col in df.columns if col not in priority_columns]
        column_order = [col for col in priority_columns if col in df.columns] + remaining_columns
        df = df[column_order]
        
        # 숫자 컬럼들의 데이터 타입 변환
        numeric_columns = [
            "판매수량", "판매가($)", "순매출액($)", "할인액($)",
            "총매출액(￦)", "순매출액(￦)", "할인액(￦)"
        ]
        
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # 날짜 컬럼 포맷팅
        date_columns = ["매출일자", "원매출일자"]
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], format='%Y%m%d', errors='coerce').dt.strftime('%Y-%m-%d')
        
        # 파일명 생성
        filename = f"신라매출데이터_{date_str}.xlsx"
        
        # Excel 파일로 저장 (스타일링 포함)
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='매출데이터', index=False)
            
            # 워크시트 가져오기
            worksheet = writer.sheets['매출데이터']
            
            # 컬럼 너비 자동 조정
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # 최대 50으로 제한
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # 헤더 스타일링
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 숫자 컬럼에 천단위 구분자 적용
            from openpyxl.styles import NamedStyle
            
            # 통화 스타일 생성
            currency_style = NamedStyle(name="currency")
            currency_style.number_format = '#,##0'
            
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in numeric_columns:
                    for row_idx in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.number_format = '#,##0'
        
        print(f"✔ Excel 파일 저장 완료: {filename}")
        print(f"📊 총 {len(df)} 건의 데이터가 저장되었습니다.")
        
        # 간단한 통계 정보
        total_sales_won = df["순매출액(￦)"].sum() if "순매출액(￦)" in df.columns else 0
        total_sales_usd = df["순매출액($)"].sum() if "순매출액($)" in df.columns else 0
        total_qty = df["판매수량"].sum() if "판매수량" in df.columns else 0
        
        print(f"💰 총 매출액: {total_sales_won:,.0f}원 / ${total_sales_usd:,.0f}")
        print(f"📦 총 판매수량: {total_qty:,.0f}개")
        
        return filename
        
    except Exception as e:
        print(f"❌ Excel 파일 저장 실패: {e}")
        import traceback
        traceback.print_exc()
        return None
    """데이터를 JSON 파일로 저장"""
    filename = f"shilla_sales_{date_str}.json"
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"✔ 데이터 파일 저장: {filename}")
    except Exception as e:
        print(f"❌ 파일 저장 실패: {e}")

def print_data_summary(data):
    """데이터 요약 정보 출력"""
    try:
        results = data.get('BODY', {}).get('t_result', [])
        if not results:
            print("❌ 매출 데이터가 없습니다.")
            return
        
        print("\n" + "="*50)
        print("📊 매출 데이터 요약")
        print("="*50)
        
        # 기본 통계
        total_records = len(results)
        total_sales = sum(float(item.get('net_amt', 0)) for item in results)
        total_qty = sum(int(item.get('qty', 0)) for item in results)
        
        print(f"📈 총 거래 건수: {total_records:,}건")
        print(f"💰 총 매출액: {total_sales:,.0f}원")
        print(f"📦 총 판매 수량: {total_qty:,}개")
        
        # 가이드별 매출
        guide_sales = {}
        for item in results:
            guide = item.get('guide_nm', '미상')
            net_amt = float(item.get('net_amt', 0))
            guide_sales[guide] = guide_sales.get(guide, 0) + net_amt
        
        print(f"\n👥 가이드별 매출:")
        for guide, sales in sorted(guide_sales.items(), key=lambda x: x[1], reverse=True):
            print(f"   {guide}: {sales:,.0f}원")
        
        # 브랜드별 매출
        brand_sales = {}
        for item in results:
            brand = item.get('brand_nm', '미상')
            net_amt = float(item.get('net_amt', 0))
            brand_sales[brand] = brand_sales.get(brand, 0) + net_amt
        
        print(f"\n🏷️ 브랜드별 매출 TOP 10:")
        for brand, sales in sorted(brand_sales.items(), key=lambda x: x[1], reverse=True)[:10]:
            print(f"   {brand}: {sales:,.0f}원")
        
        # 카테고리별 매출
        category_sales = {}
        for item in results:
            category = item.get('cat_nm', '미상')
            net_amt = float(item.get('net_amt', 0))
            category_sales[category] = category_sales.get(category, 0) + net_amt
        
        print(f"\n📂 카테고리별 매출:")
        for category, sales in sorted(category_sales.items(), key=lambda x: x[1], reverse=True):
            print(f"   {category}: {sales:,.0f}원")
        
        # 매출 유형별 분석
        sale_type_sales = {}
        for item in results:
            sale_type = item.get('sale_fg_nm', '미상')
            net_amt = float(item.get('net_amt', 0))
            sale_type_sales[sale_type] = sale_type_sales.get(sale_type, 0) + net_amt
        
        print(f"\n🛒 매출 유형별:")
        for sale_type, sales in sorted(sale_type_sales.items(), key=lambda x: x[1], reverse=True):
            print(f"   {sale_type}: {sales:,.0f}원")
        
        print("="*50)
        
    except Exception as e:
        print(f"❌ 데이터 요약 생성 실패: {e}")

def analyze_specific_guide(data, guide_name):
    """특정 가이드의 매출 분석"""
    try:
        results = data.get('BODY', {}).get('t_result', [])
        guide_data = [item for item in results if item.get('guide_nm') == guide_name]
        
        if not guide_data:
            print(f"❌ '{guide_name}' 가이드의 데이터가 없습니다.")
            return
        
        print(f"\n🔍 {guide_name} 가이드 상세 분석")
        print("-" * 40)
        
        total_sales = sum(float(item.get('net_amt', 0)) for item in guide_data)
        total_qty = sum(int(item.get('qty', 0)) for item in guide_data)
        
        print(f"💰 총 매출: {total_sales:,.0f}원")
        print(f"📦 총 수량: {total_qty:,}개")
        print(f"📊 거래 건수: {len(guide_data):,}건")
        
        # 브랜드별 매출
        brand_sales = {}
        for item in guide_data:
            brand = item.get('brand_nm', '미상')
            net_amt = float(item.get('net_amt', 0))
            brand_sales[brand] = brand_sales.get(brand, 0) + net_amt
        
        print(f"\n🏷️ 브랜드별 매출:")
        for brand, sales in sorted(brand_sales.items(), key=lambda x: x[1], reverse=True):
            print(f"   {brand}: {sales:,.0f}원")
        
    except Exception as e:
        print(f"❌ 가이드 분석 실패: {e}")

if __name__ == "__main__":
    # 네트워크 연결 확인
    if not check_network_connection():
        print("❌ 네트워크 연결을 확인해주세요.")
        exit(1)
    
    # 기본 매출 데이터 가져오기
    result = automate_shilla_download()
    
    if result:
        # 특정 가이드 분석 (예시)
        # analyze_specific_guide(result, "원미화")
        
        # 추가 분석이 필요한 경우 여기에 코드 추가
        pass